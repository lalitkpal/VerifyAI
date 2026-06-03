import csv
import io
import openpyxl
from typing import List, Dict, Any

def parse_csv_data(content: str) -> List[Dict[str, Any]]:
    """
    Parses CSV content and extracts test cases.
    Looks for headers like prompt, message, and expected_output.
    """
    f = io.StringIO(content)
    reader = csv.DictReader(f)
    headers = reader.fieldnames
    if not headers:
        # Fallback if empty
        return []
        
    # Helper to find header match
    def find_header(keys: List[str], default: str) -> str:
        for k in keys:
            for h in headers:
                if h.lower().strip() == k:
                    return h
        return default

    prompt_col = find_header(["prompt", "input", "question"], "prompt")
    message_col = find_header(["message", "output", "response", "generated_text", "generated"], "message")
    expected_col = find_header(["expected_output", "expected", "reference", "target"], "expected_output")

    rows = []
    for row in reader:
        prompt = row.get(prompt_col, "").strip() if prompt_col in row and row[prompt_col] else ""
        message = row.get(message_col, "").strip() if message_col in row and row[message_col] else ""
        expected = row.get(expected_col, "").strip() if expected_col in row and row[expected_col] else ""
        
        # Fallback to column position if column mapping wasn't found but columns exist
        vals = list(row.values())
        if not prompt and len(vals) > 0:
            prompt = str(vals[0] or "").strip()
        if not message and len(vals) > 1:
            message = str(vals[1] or "").strip()
        if not expected and len(vals) > 2:
            expected = str(vals[2] or "").strip()
            
        if prompt and message:
            rows.append({
                "prompt": prompt,
                "message": message,
                "expected_output": expected or None
            })
    return rows

def parse_excel_data(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parses Excel (.xlsx/.xls) file bytes using openpyxl and extracts test cases.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = []
    
    if sheet.max_row < 1:
        return []

    # Read the headers from the first row
    headers = []
    for cell in sheet[1]:
        val = str(cell.value or "").strip().lower()
        headers.append(val)

    def find_col_idx(keys: List[str]) -> int:
        for k in keys:
            for idx, h in enumerate(headers):
                if h == k:
                    return idx
        return -1

    prompt_idx = find_col_idx(["prompt", "input", "question"])
    message_idx = find_col_idx(["message", "output", "response", "generated_text", "generated"])
    expected_idx = find_col_idx(["expected_output", "expected", "reference", "target"])

    # Fallbacks if index not found
    if prompt_idx == -1:
        prompt_idx = 0
    if message_idx == -1:
        message_idx = 1 if len(headers) > 1 else -1
    if expected_idx == -1:
        expected_idx = 2 if len(headers) > 2 else -1

    for r in range(2, sheet.max_row + 1):
        row_vals = []
        for c in range(1, sheet.max_column + 1):
            row_vals.append(sheet.cell(row=r, column=c).value)
            
        if not row_vals or all(v is None for v in row_vals):
            continue

        prompt = str(row_vals[prompt_idx]).strip() if prompt_idx < len(row_vals) and row_vals[prompt_idx] is not None else ""
        message = str(row_vals[message_idx]).strip() if message_idx != -1 and message_idx < len(row_vals) and row_vals[message_idx] is not None else ""
        expected = str(row_vals[expected_idx]).strip() if expected_idx != -1 and expected_idx < len(row_vals) and row_vals[expected_idx] is not None else ""

        if prompt and message:
            rows.append({
                "prompt": prompt,
                "message": message,
                "expected_output": expected or None
            })
            
    return rows
